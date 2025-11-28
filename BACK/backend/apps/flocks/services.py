from __future__ import annotations

import datetime
from typing import List

from django.conf import settings
from django.db import transaction
from django.utils import timezone

from openpyxl import load_workbook

from .models import BreedReference, ReferenceImportLog


class BreedReferenceService:
    """Servicios relacionados con la tabla BreedReference, incluyendo import desde Excel."""

    @staticmethod
    def import_from_excel(file_path: str, imported_by) -> ReferenceImportLog:
        """Importa una hoja de Excel con columnas esperadas y crea/actualiza referencias.

        Columnas requeridas: breed, age_days, expected_weight
        Opcionales: expected_consumption, tolerance_range

        Reglas:
        - Para cada fila válida se crea una nueva versión (version = max_version + 1) y se marca is_active=True.
        - Las versiones previas para la misma (breed, age_days) se desactivan.
        - Registra el resultado en ReferenceImportLog con conteo de éxitos y errores.
        """

        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb.active

        header = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        required = ['breed', 'age_days', 'expected_weight']
        col_map = {}
        for idx, name in enumerate(header):
            if name:
                col_map[name] = idx + 1

        missing = [c for c in required if c not in col_map]
        log = ReferenceImportLog.objects.create(
            file_name=file_path.split('/')[-1] if '/' in file_path else file_path.split('\\')[-1],
            imported_by=imported_by,
            total_rows=0,
            successful_imports=0,
            updates=0,
            errors=0,
            error_details=[],
        )

        if missing:
            log.errors = 0
            log.error_details = [f"missing columns: {missing}"]
            log.save()
            return log

        total = 0
        successes = 0
        updates = 0
        errors = 0
        error_details: List[str] = []

        for row in ws.iter_rows(min_row=2):
            total += 1
            try:
                breed = row[col_map['breed'] - 1].value
                age_days = row[col_map['age_days'] - 1].value
                expected_weight = row[col_map['expected_weight'] - 1].value

                if breed is None or age_days is None or expected_weight is None:
                    raise ValueError('required field missing')

                expected_consumption = None
                if 'expected_consumption' in col_map:
                    expected_consumption = row[col_map['expected_consumption'] - 1].value or 0

                tolerance_range = None
                if 'tolerance_range' in col_map:
                    tolerance_range = row[col_map['tolerance_range'] - 1].value or 10.0

                with transaction.atomic():
                    # determine new version
                    latest = BreedReference.objects.filter(breed=breed, age_days=int(age_days)).order_by('-version').first()
                    new_version = 1
                    if latest:
                        new_version = latest.version + 1
                        # deactivate previous versions
                        BreedReference.objects.filter(breed=breed, age_days=int(age_days), is_active=True).update(is_active=False)

                    br = BreedReference.objects.create(
                        breed=str(breed),
                        age_days=int(age_days),
                        expected_weight=float(expected_weight),
                        expected_consumption=float(expected_consumption or 0),
                        tolerance_range=float(tolerance_range or 10.0),
                        version=new_version,
                        is_active=True,
                        created_by=imported_by,
                    )

                successes += 1
            except Exception as exc:  # pragma: no cover - exception paths logged
                errors += 1
                error_details.append(f'row {total+1}: {str(exc)}')

        log.total_rows = total
        log.successful_imports = successes
        log.updates = updates
        log.errors = errors
        log.error_details = error_details
        log.save()

        return log

from django.db import transaction, models
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.core.exceptions import ValidationError

from .models import MortalityRecord, MortalityCause, Flock


class MortalityService:
    @staticmethod
    def register_mortality_batch(mortality_records, user):
        """Registra múltiples mortalidades en batch (para sync offline)"""
        results = []

        with transaction.atomic():
            for record_data in mortality_records:
                try:
                    result = MortalityService._process_single_mortality(record_data, user)
                    results.append(result)
                except Exception as e:
                    results.append({
                        'client_id': record_data.get('client_id'),
                        'status': 'error',
                        'error': str(e)
                    })

        return results

    @staticmethod
    def _process_single_mortality(record_data, user):
        flock = Flock.objects.get(id=record_data['flock_id'])
        date = parse_date(record_data['date'])
        deaths = int(record_data['deaths'])

        # Validar permisos (Galponero solo en sus galpones)
        role_name = getattr(getattr(user, 'role', None), 'name', None)
        if role_name == 'Galponero' and flock.shed.assigned_worker != user:
            raise PermissionError("No tienes permisos para registrar mortalidad en este galpón")

        # Verificar duplicado del mismo día (sumar si existe)
        existing = MortalityRecord.objects.filter(flock=flock, date=date).first()

        if existing:
            # Sumar mortalidad al registro existente
            total_deaths = existing.deaths + deaths
            if total_deaths > (flock.current_quantity + existing.deaths):
                raise ValidationError("La mortalidad total excede la cantidad del lote")

            existing.deaths = total_deaths
            existing.save()  # Trigger actualización automática

            return {
                'client_id': record_data.get('client_id'),
                'status': 'success',
                'action': 'updated',
                'server_id': existing.id
            }

        # Crear nuevo registro
        cause = None
        if record_data.get('cause_name'):
            cause, _ = MortalityCause.objects.get_or_create(
                name=record_data['cause_name'],
                defaults={'category': 'OTHER'}
            )

        mortality_record = MortalityRecord.objects.create(
            flock=flock,
            date=date,
            deaths=deaths,
            cause=cause,
            temperature=record_data.get('temperature'),
            notes=record_data.get('notes', ''),
            recorded_by=user,
            client_id=record_data.get('client_id'),
            created_by_device=record_data.get('device_id')
        )

        return {
            'client_id': record_data.get('client_id'),
            'status': 'success',
            'action': 'created',
            'server_id': mortality_record.id
        }

    @staticmethod
    def calculate_mortality_stats(flock, days=7):
        """Calcula estadísticas de mortalidad de un lote"""
        from datetime import timedelta
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days - 1)

        # Get records in the range
        mortality_records = flock.mortality_records.filter(date__range=[start_date, end_date])

        total_deaths = mortality_records.aggregate(total=models.Sum('deaths'))['total'] or 0

        # Build a daily series (one entry per day in the range)
        series = []
        for i in range(days):
            day = start_date + timedelta(days=i)
            day_deaths = mortality_records.filter(date=day).aggregate(total=models.Sum('deaths'))['total'] or 0
            day_rate = (day_deaths / flock.initial_quantity) * 100 if flock.initial_quantity else 0
            series.append({
                'date': day.isoformat(),
                'label': day.strftime('%Y-%m-%d'),
                'deaths': int(day_deaths),
                'mortality_rate': float(day_rate),
                'industry_average': None,
            })

        return {
            'total_deaths': int(total_deaths),
            'mortality_rate': float((total_deaths / flock.initial_quantity) * 100) if flock.initial_quantity else 0.0,
            'daily_average': float(total_deaths / days) if days else 0.0,
            'worst_day': mortality_records.order_by('-deaths').first(),
            'period': f'{start_date} - {end_date}',
            'series': series,
        }
