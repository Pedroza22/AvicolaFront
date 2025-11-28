from django.db.models import Avg, Sum, Count, Q, F, Max, Min
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from typing import Dict, List, Optional, Any

from apps.flocks.models import Flock, DailyWeightRecord, MortalityRecord
from apps.inventory.models import InventoryItem, FoodConsumptionRecord
from apps.alarms.models import Alarm
from apps.farms.models import Farm, Shed
from .models import Report, ReportStatus


class ProductivityReportService:
    """Servicio para generar reportes de productividad"""
    
    def __init__(self, report: Report):
        self.report = report
        self.farm = report.farm
        self.shed = report.shed
        self.flock = report.flock
        self.date_from = report.date_from
        self.date_to = report.date_to
    
    def generate_report(self) -> Dict[str, Any]:
        """Genera el reporte completo de productividad"""
        try:
            self.report.set_processing()
            
            # Obtener datos base
            flocks = self._get_flocks()
            
            report_data = {
                'report_info': {
                    'name': self.report.name,
                    'type': 'productivity',
                    'period': {
                        'from': self.date_from.isoformat(),
                        'to': self.date_to.isoformat(),
                        'days': self.report.duration_days
                    },
                    'scope': {
                        'farm': self.farm.name if self.farm else 'Todas las fincas',
                        'shed': self.shed.name if self.shed else 'Todos los galpones',
                        'flock': self.flock.name if self.flock else 'Todos los lotes',
                        'flocks_analyzed': flocks.count()
                    },
                    'generated_at': timezone.now().isoformat()
                },
                'summary': self._generate_summary(flocks),
                'weight_analysis': self._analyze_weight_performance(flocks),
                'mortality_analysis': self._analyze_mortality(flocks),
                'consumption_analysis': self._analyze_consumption(flocks),
                'conversion_analysis': self._analyze_feed_conversion(flocks),
                'comparative_analysis': self._generate_comparative_analysis(flocks),
                'trends': self._analyze_trends(flocks),
                'alerts': self._generate_alerts(flocks)
            }
            
            # Generar archivo Excel si se requiere
            if self.report.export_format == 'excel':
                file_path = self._generate_excel_report(report_data)
                self.report.set_completed(report_data, file_path)
            else:
                self.report.set_completed(report_data)
            
            return report_data
            
        except Exception as e:
            self.report.set_failed(str(e))
            raise
    
    def _get_flocks(self):
        """Obtiene los lotes según los filtros aplicados"""
        queryset = Flock.objects.all()
        
        if self.flock:
            return queryset.filter(id=self.flock.id)
        
        if self.shed:
            queryset = queryset.filter(shed=self.shed)
        elif self.farm:
            queryset = queryset.filter(shed__farm=self.farm)
        
        # Solo lotes que tengan datos en el período
        queryset = queryset.filter(
            Q(placement_date__lte=self.date_to) &
            (Q(sale_date__gte=self.date_from) | Q(sale_date__isnull=True))
        )
        
        return queryset.select_related('shed', 'shed__farm', 'breed')
    
    def _generate_summary(self, flocks) -> Dict[str, Any]:
        """Genera resumen ejecutivo"""
        total_birds = flocks.aggregate(Sum('initial_quantity'))['initial_quantity__sum'] or 0
        active_flocks = flocks.filter(sale_date__isnull=True).count()
        
        # Mortalidad total en el período
        mortality_data = MortalityRecord.objects.filter(
            flock__in=flocks,
            date__range=[self.date_from, self.date_to]
        ).aggregate(
            total_deaths=Sum('quantity'),
            records_count=Count('id')
        )
        
        # Peso promedio actual
        latest_weights = DailyWeightRecord.objects.filter(
            flock__in=flocks,
            date__range=[self.date_from, self.date_to]
        ).aggregate(
            avg_weight=Avg('average_weight'),
            records_count=Count('id')
        )
        
        # Consumo total
        consumption_data = FoodConsumptionRecord.objects.filter(
            flock__in=flocks,
            consumption_date__range=[self.date_from, self.date_to]
        ).aggregate(
            total_consumption=Sum('quantity_kg'),
            records_count=Count('id')
        )
        
        return {
            'total_flocks': flocks.count(),
            'active_flocks': active_flocks,
            'total_birds': total_birds,
            'mortality': {
                'total_deaths': mortality_data['total_deaths'] or 0,
                'mortality_rate': round((mortality_data['total_deaths'] or 0) / max(total_birds, 1) * 100, 2),
                'records': mortality_data['records_count'] or 0
            },
            'weight': {
                'average_weight': round(latest_weights['avg_weight'] or 0, 2),
                'records': latest_weights['records_count'] or 0
            },
            'consumption': {
                'total_kg': round(consumption_data['total_consumption'] or 0, 2),
                'records': consumption_data['records_count'] or 0
            }
        }
    
    def _analyze_weight_performance(self, flocks) -> Dict[str, Any]:
        """Analiza el rendimiento de peso"""
        weight_records = DailyWeightRecord.objects.filter(
            flock__in=flocks,
            date__range=[self.date_from, self.date_to]
        ).select_related('flock', 'flock__breed')
        
        # Análisis por lote
        flock_analysis = []
        
        for flock in flocks:
            flock_weights = weight_records.filter(flock=flock).order_by('date')
            
            if flock_weights.exists():
                first_weight = flock_weights.first()
                last_weight = flock_weights.last()
                avg_weight = flock_weights.aggregate(Avg('average_weight'))['average_weight__avg']
                
                # Calcular ganancia diaria promedio
                if flock_weights.count() > 1:
                    weight_gain = (last_weight.average_weight - first_weight.average_weight) / max((last_weight.date - first_weight.date).days, 1)
                else:
                    weight_gain = 0
                
                # Comparar con estándar de la raza
                breed_standard = None
                deviation = None
                if flock.breed and hasattr(flock.breed, 'weight_standards'):
                    age_days = (timezone.now().date() - flock.placement_date).days
                    breed_standard = flock.breed.get_expected_weight(age_days)
                    if breed_standard and avg_weight:
                        deviation = ((avg_weight - breed_standard) / breed_standard) * 100
                
                flock_analysis.append({
                    'flock_id': flock.id,
                    'flock_name': flock.name,
                    'breed': flock.breed.name if flock.breed else None,
                    'age_days': (timezone.now().date() - flock.placement_date).days,
                    'records_count': flock_weights.count(),
                    'weights': {
                        'first': round(first_weight.average_weight, 2),
                        'last': round(last_weight.average_weight, 2),
                        'average': round(avg_weight, 2),
                        'daily_gain': round(weight_gain, 3)
                    },
                    'comparison': {
                        'breed_standard': round(breed_standard, 2) if breed_standard else None,
                        'deviation_percent': round(deviation, 2) if deviation else None,
                        'performance': 'above' if deviation and deviation > 5 else 'below' if deviation and deviation < -5 else 'normal'
                    }
                })
        
        # Tendencias generales
        daily_weights = weight_records.values('date').annotate(
            avg_weight=Avg('average_weight'),
            records_count=Count('id')
        ).order_by('date')
        
        return {
            'flock_analysis': flock_analysis,
            'daily_trends': list(daily_weights),
            'overall_performance': {
                'best_performer': max(flock_analysis, key=lambda x: x['weights']['daily_gain']) if flock_analysis else None,
                'worst_performer': min(flock_analysis, key=lambda x: x['weights']['daily_gain']) if flock_analysis else None,
                'average_daily_gain': sum(f['weights']['daily_gain'] for f in flock_analysis) / len(flock_analysis) if flock_analysis else 0
            }
        }
    
    def _analyze_mortality(self, flocks) -> Dict[str, Any]:
        """Analiza la mortalidad"""
        mortality_records = MortalityRecord.objects.filter(
            flock__in=flocks,
            date__range=[self.date_from, self.date_to]
        ).select_related('flock')
        
        # Análisis por causa
        by_cause = mortality_records.values('cause').annotate(
            total_deaths=Sum('quantity'),
            records_count=Count('id')
        ).order_by('-total_deaths')
        
        # Análisis por lote
        by_flock = mortality_records.values('flock__name', 'flock__id').annotate(
            total_deaths=Sum('quantity'),
            mortality_rate=Sum('quantity') * 100.0 / F('flock__initial_quantity')
        ).order_by('-total_deaths')
        
        # Tendencia diaria
        daily_mortality = mortality_records.values('date').annotate(
            total_deaths=Sum('quantity'),
            records_count=Count('id')
        ).order_by('date')
        
        return {
            'total_deaths': mortality_records.aggregate(Sum('quantity'))['quantity__sum'] or 0,
            'total_records': mortality_records.count(),
            'by_cause': list(by_cause),
            'by_flock': list(by_flock),
            'daily_trends': list(daily_mortality),
            'highest_mortality_day': max(daily_mortality, key=lambda x: x['total_deaths']) if daily_mortality else None
        }
    
    def _analyze_consumption(self, flocks) -> Dict[str, Any]:
        """Analiza el consumo de alimento"""
        consumption_records = FoodConsumptionRecord.objects.filter(
            flock__in=flocks,
            consumption_date__range=[self.date_from, self.date_to]
        ).select_related('flock', 'food_batch', 'food_batch__inventory_item')
        
        # Consumo por lote
        by_flock = consumption_records.values('flock__name', 'flock__id').annotate(
            total_consumption=Sum('quantity_kg'),
            avg_daily_consumption=Avg('quantity_kg'),
            records_count=Count('id')
        ).order_by('-total_consumption')
        
        # Consumo por tipo de alimento
        by_food_type = consumption_records.values('food_batch__inventory_item__name').annotate(
            total_consumption=Sum('quantity_kg'),
            records_count=Count('id')
        ).order_by('-total_consumption')
        
        # Tendencia diaria
        daily_consumption = consumption_records.values('consumption_date').annotate(
            total_consumption=Sum('quantity_kg'),
            records_count=Count('id')
        ).order_by('consumption_date')
        
        return {
            'total_consumption': consumption_records.aggregate(Sum('quantity_kg'))['quantity_kg__sum'] or 0,
            'by_flock': list(by_flock),
            'by_food_type': list(by_food_type),
            'daily_trends': list(daily_consumption),
            'average_daily': sum(d['total_consumption'] for d in daily_consumption) / len(daily_consumption) if daily_consumption else 0
        }
    
    def _analyze_feed_conversion(self, flocks) -> Dict[str, Any]:
        """Analiza la conversión alimenticia"""
        conversion_data = []
        
        for flock in flocks:
            # Consumo total del lote en el período
            total_consumption = FoodConsumptionRecord.objects.filter(
                flock=flock,
                consumption_date__range=[self.date_from, self.date_to]
            ).aggregate(Sum('quantity_kg'))['quantity_kg__sum'] or 0
            
            # Ganancia de peso en el período
            weight_records = DailyWeightRecord.objects.filter(
                flock=flock,
                date__range=[self.date_from, self.date_to]
            ).order_by('date')
            
            if weight_records.count() >= 2:
                first_weight = weight_records.first().average_weight
                last_weight = weight_records.last().average_weight
                weight_gain_per_bird = last_weight - first_weight
                total_weight_gain = weight_gain_per_bird * flock.current_quantity
                
                # Conversión alimenticia (kg alimento / kg ganancia)
                feed_conversion = total_consumption / max(total_weight_gain, 0.001)
                
                conversion_data.append({
                    'flock_id': flock.id,
                    'flock_name': flock.name,
                    'total_consumption_kg': round(total_consumption, 2),
                    'total_weight_gain_kg': round(total_weight_gain, 2),
                    'feed_conversion_ratio': round(feed_conversion, 3),
                    'efficiency': 'excellent' if feed_conversion < 1.8 else 'good' if feed_conversion < 2.2 else 'poor'
                })
        
        return {
            'flock_conversions': conversion_data,
            'average_conversion': sum(f['feed_conversion_ratio'] for f in conversion_data) / len(conversion_data) if conversion_data else 0,
            'best_converter': min(conversion_data, key=lambda x: x['feed_conversion_ratio']) if conversion_data else None,
            'worst_converter': max(conversion_data, key=lambda x: x['feed_conversion_ratio']) if conversion_data else None
        }
    
    def _generate_comparative_analysis(self, flocks) -> Dict[str, Any]:
        """Genera análisis comparativo con períodos anteriores"""
        # Período anterior del mismo tamaño
        days_diff = (self.date_to - self.date_from).days
        previous_date_to = self.date_from - timedelta(days=1)
        previous_date_from = previous_date_to - timedelta(days=days_diff)
        
        current_data = self._get_period_metrics(flocks, self.date_from, self.date_to)
        previous_data = self._get_period_metrics(flocks, previous_date_from, previous_date_to)
        
        return {
            'current_period': current_data,
            'previous_period': previous_data,
            'comparison': {
                'mortality_change': self._calculate_change(current_data['mortality_rate'], previous_data['mortality_rate']),
                'weight_change': self._calculate_change(current_data['avg_weight'], previous_data['avg_weight']),
                'consumption_change': self._calculate_change(current_data['total_consumption'], previous_data['total_consumption'])
            }
        }
    
    def _get_period_metrics(self, flocks, date_from, date_to) -> Dict[str, Any]:
        """Obtiene métricas para un período específico"""
        mortality = MortalityRecord.objects.filter(
            flock__in=flocks,
            date__range=[date_from, date_to]
        ).aggregate(Sum('quantity'))['quantity__sum'] or 0
        
        weight = DailyWeightRecord.objects.filter(
            flock__in=flocks,
            date__range=[date_from, date_to]
        ).aggregate(Avg('average_weight'))['average_weight__avg'] or 0
        
        consumption = FoodConsumptionRecord.objects.filter(
            flock__in=flocks,
            consumption_date__range=[date_from, date_to]
        ).aggregate(Sum('quantity_kg'))['quantity_kg__sum'] or 0
        
        total_birds = flocks.aggregate(Sum('initial_quantity'))['initial_quantity__sum'] or 1
        
        return {
            'mortality_rate': (mortality / total_birds) * 100,
            'avg_weight': weight,
            'total_consumption': consumption
        }
    
    def _calculate_change(self, current, previous) -> Dict[str, Any]:
        """Calcula el cambio porcentual entre dos valores"""
        if previous == 0:
            return {'value': 0, 'percent': 0, 'trend': 'stable'}
        
        change_percent = ((current - previous) / previous) * 100
        
        return {
            'value': round(current - previous, 2),
            'percent': round(change_percent, 2),
            'trend': 'up' if change_percent > 5 else 'down' if change_percent < -5 else 'stable'
        }
    
    def _analyze_trends(self, flocks) -> Dict[str, Any]:
        """Analiza tendencias en el tiempo"""
        # Implementar análisis de tendencias
        return {
            'weight_trend': 'increasing',  # Simplificado por ahora
            'mortality_trend': 'stable',
            'consumption_trend': 'stable'
        }
    
    def _generate_alerts(self, flocks) -> List[Dict[str, Any]]:
        """Genera alertas basadas en el análisis"""
        alerts = []
        
        # Verificar alertas de mortalidad alta
        for flock in flocks:
            mortality_rate = MortalityRecord.objects.filter(
                flock=flock,
                date__range=[self.date_from, self.date_to]
            ).aggregate(Sum('quantity'))['quantity__sum'] or 0
            
            mortality_percent = (mortality_rate / flock.initial_quantity) * 100
            
            if mortality_percent > 5:  # Más del 5% de mortalidad
                alerts.append({
                    'type': 'high_mortality',
                    'severity': 'high',
                    'flock': flock.name,
                    'message': f'Mortalidad alta: {mortality_percent:.1f}% en {flock.name}',
                    'value': mortality_percent
                })
        
        return alerts
    
    def _generate_excel_report(self, data: Dict[str, Any]) -> str:
        """Genera archivo Excel con gráficos"""
        # Crear directorio si no existe
        reports_dir = os.path.join('media', 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        # Nombre del archivo
        filename = f"productivity_report_{self.report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(reports_dir, filename)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Hoja de resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen Ejecutivo"
        self._create_summary_sheet(ws_summary, data)
        
        # Hoja de análisis de peso
        ws_weight = wb.create_sheet("Análisis de Peso")
        self._create_weight_sheet(ws_weight, data['weight_analysis'])
        
        # Hoja de mortalidad
        ws_mortality = wb.create_sheet("Mortalidad")
        self._create_mortality_sheet(ws_mortality, data['mortality_analysis'])
        
        # Hoja de consumo
        ws_consumption = wb.create_sheet("Consumo")
        self._create_consumption_sheet(ws_consumption, data['consumption_analysis'])
        
        wb.save(filepath)
        return filepath
    
    def _create_summary_sheet(self, ws, data):
        """Crea la hoja de resumen ejecutivo"""
        # Título
        ws['A1'] = data['report_info']['name']
        ws['A1'].font = Font(size=16, bold=True)
        
        # Información del período
        ws['A3'] = "Período de Análisis:"
        ws['B3'] = f"{data['report_info']['period']['from']} a {data['report_info']['period']['to']}"
        
        # Métricas principales
        row = 5
        summary = data['summary']
        
        metrics = [
            ("Total de Lotes", summary['total_flocks']),
            ("Lotes Activos", summary['active_flocks']),
            ("Total de Aves", summary['total_birds']),
            ("Mortalidad Total", summary['mortality']['total_deaths']),
            ("Tasa de Mortalidad", f"{summary['mortality']['mortality_rate']}%"),
            ("Peso Promedio", f"{summary['weight']['average_weight']} kg"),
            ("Consumo Total", f"{summary['consumption']['total_kg']} kg")
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
    
    def _create_weight_sheet(self, ws, weight_data):
        """Crea la hoja de análisis de peso"""
        ws['A1'] = "Análisis de Peso por Lote"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ["Lote", "Raza", "Edad (días)", "Peso Inicial", "Peso Final", "Peso Promedio", "Ganancia Diaria"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).font = Font(bold=True)
        
        # Datos
        for row, flock_data in enumerate(weight_data['flock_analysis'], 4):
            ws.cell(row=row, column=1, value=flock_data['flock_name'])
            ws.cell(row=row, column=2, value=flock_data['breed'] or 'N/A')
            ws.cell(row=row, column=3, value=flock_data['age_days'])
            ws.cell(row=row, column=4, value=flock_data['weights']['first'])
            ws.cell(row=row, column=5, value=flock_data['weights']['last'])
            ws.cell(row=row, column=6, value=flock_data['weights']['average'])
            ws.cell(row=row, column=7, value=flock_data['weights']['daily_gain'])
    
    def _create_mortality_sheet(self, ws, mortality_data):
        """Crea la hoja de análisis de mortalidad"""
        ws['A1'] = "Análisis de Mortalidad"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Resumen
        ws['A3'] = "Total de Muertes:"
        ws['B3'] = mortality_data['total_deaths']
        
        # Por causa
        ws['A6'] = "Mortalidad por Causa"
        ws['A6'].font = Font(bold=True)
        
        headers = ["Causa", "Muertes", "Registros"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header).font = Font(bold=True)
        
        for row, cause_data in enumerate(mortality_data['by_cause'], 8):
            ws.cell(row=row, column=1, value=cause_data['cause'])
            ws.cell(row=row, column=2, value=cause_data['total_deaths'])
            ws.cell(row=row, column=3, value=cause_data['records_count'])
    
    def _create_consumption_sheet(self, ws, consumption_data):
        """Crea la hoja de análisis de consumo"""
        ws['A1'] = "Análisis de Consumo"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Resumen
        ws['A3'] = "Consumo Total:"
        ws['B3'] = f"{consumption_data['total_consumption']} kg"
        
        # Por lote
        ws['A6'] = "Consumo por Lote"
        ws['A6'].font = Font(bold=True)
        
        headers = ["Lote", "Consumo Total (kg)", "Consumo Promedio Diario", "Registros"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header).font = Font(bold=True)
        
        for row, flock_data in enumerate(consumption_data['by_flock'], 8):
            ws.cell(row=row, column=1, value=flock_data['flock__name'])
            ws.cell(row=row, column=2, value=round(flock_data['total_consumption'], 2))
            ws.cell(row=row, column=3, value=round(flock_data['avg_daily_consumption'], 2))
            ws.cell(row=row, column=4, value=flock_data['records_count'])